import os
import json
import tempfile
import shutil
from pathlib import Path
from typing import Dict, Optional, Tuple, List
import traceback

# Importaciones condicionales para diferentes tipos de archivos
try:
    from docx import Document as DocxDocument
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import textract
    TEXTRACT_AVAILABLE = True
except ImportError:
    TEXTRACT_AVAILABLE = False

try:
    import docx2txt
    DOCX2TXT_AVAILABLE = True
except ImportError:
    DOCX2TXT_AVAILABLE = False

try:
    import pypandoc
    PYPANDOC_AVAILABLE = True
except ImportError:
    PYPANDOC_AVAILABLE = False

try:
    import PyPDF2
    from pdfplumber import PDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class ProcessorUniversal:
    """Procesador universal de documentos legales."""
    
    def __init__(self):
        self.supported_formats = {
            '.docx': self._process_docx,
            '.doc': self._process_doc,
            '.pdf': self._process_pdf,
            '.txt': self._process_txt,
            '.rtf': self._process_rtf,
            '.odt': self._process_odt,
            '.xlsx': self._process_excel,
            '.xls': self._process_excel,
        }
        
        self._check_dependencies()
    
    def _check_dependencies(self):
        """Verifica qué dependencias están disponibles."""
        print("🔍 VERIFICANDO DEPENDENCIAS:")
        deps = {
            'python-docx (.docx)': DOCX_AVAILABLE,
            'textract (universal)': TEXTRACT_AVAILABLE,
            'docx2txt (.doc)': DOCX2TXT_AVAILABLE,
            'pypandoc (conversión)': PYPANDOC_AVAILABLE,
            'PyPDF2 (PDF)': PDF_AVAILABLE,
            'openpyxl (Excel)': EXCEL_AVAILABLE,
        }
        
        for dep, available in deps.items():
            status = "✅" if available else "❌"
            print(f"  {status} {dep}")
        print()
    
    def extract_text_from_file(self, filepath: str) -> Tuple[str, str]:
        """
        Extrae texto de cualquier tipo de archivo soportado.
        Returns: (texto_extraido, metodo_usado)
        """
        file_ext = Path(filepath).suffix.lower()
        filename = os.path.basename(filepath)
        
        print(f"📄 Procesando: {filename} ({file_ext})")
        
        # Intentar método específico para la extensión
        if file_ext in self.supported_formats:
            try:
                text, method = self.supported_formats[file_ext](filepath)
                if text.strip():
                    print(f"  ✅ Exitoso con método específico: {method}")
                    return text, method
            except Exception as e:
                print(f"  ⚠️ Método específico falló: {str(e)[:100]}")
        
        # Intentar métodos universales como fallback
        fallback_methods = [
            ('textract', self._try_textract),
            ('pypandoc', self._try_pypandoc),
            ('raw_text', self._try_raw_text)
        ]
        
        for method_name, method_func in fallback_methods:
            try:
                text = method_func(filepath)
                if text and text.strip():
                    print(f"  ✅ Exitoso con fallback: {method_name}")
                    return text, f"fallback_{method_name}"
            except Exception as e:
                print(f"  ⚠️ {method_name} falló: {str(e)[:50]}")
        
        print(f"  ❌ No se pudo extraer texto")
        return "", "failed"
    
    def _process_docx(self, filepath: str) -> Tuple[str, str]:
        """Procesa archivos .docx"""
        if not DOCX_AVAILABLE:
            raise ImportError("python-docx no disponible")
        
        doc = DocxDocument(filepath)
        text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
        return text, "python-docx"
    
    def _process_doc(self, filepath: str) -> Tuple[str, str]:
        """Procesa archivos .doc antiguos"""
        # Método 1: docx2txt
        if DOCX2TXT_AVAILABLE:
            try:
                text = docx2txt.process(filepath)
                if text.strip():
                    return text, "docx2txt"
            except:
                pass
        
        # Método 2: textract
        if TEXTRACT_AVAILABLE:
            text = textract.process(filepath).decode('utf-8')
            return text, "textract"
        
        raise ImportError("No hay librerías disponibles para .doc")
    
    def _process_pdf(self, filepath: str) -> Tuple[str, str]:
        """Procesa archivos PDF"""
        if not PDF_AVAILABLE:
            raise ImportError("Librerías PDF no disponibles")
        
        # Método 1: pdfplumber (mejor para texto estructurado)
        try:
            with PDF.open(filepath) as pdf:
                text = ""
                for page in pdf.pages:
                    text += page.extract_text() + "\n"
                if text.strip():
                    return text, "pdfplumber"
        except:
            pass
        
        # Método 2: PyPDF2
        with open(filepath, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text, "PyPDF2"
    
    def _process_txt(self, filepath: str) -> Tuple[str, str]:
        """Procesa archivos de texto plano"""
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        
        for encoding in encodings:
            try:
                with open(filepath, 'r', encoding=encoding) as f:
                    return f.read(), f"txt_{encoding}"
            except UnicodeDecodeError:
                continue
        
        raise UnicodeDecodeError("No se pudo decodificar el archivo")
    
    def _process_rtf(self, filepath: str) -> Tuple[str, str]:
        """Procesa archivos RTF"""
        if PYPANDOC_AVAILABLE:
            text = pypandoc.convert_file(filepath, 'plain')
            return text, "pypandoc_rtf"
        elif TEXTRACT_AVAILABLE:
            text = textract.process(filepath).decode('utf-8')
            return text, "textract_rtf"
        else:
            raise ImportError("No hay librerías para RTF")
    
    def _process_odt(self, filepath: str) -> Tuple[str, str]:
        """Procesa archivos ODT (LibreOffice)"""
        if TEXTRACT_AVAILABLE:
            text = textract.process(filepath).decode('utf-8')
            return text, "textract_odt"
        else:
            raise ImportError("textract no disponible para ODT")
    
    def _process_excel(self, filepath: str) -> Tuple[str, str]:
        """Procesa archivos Excel"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl no disponible")
        
        wb = load_workbook(filepath, data_only=True)
        text = ""
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text += f"--- Hoja: {sheet_name} ---\n"
            
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
        
        return text, "openpyxl"
    
    def _try_textract(self, filepath: str) -> str:
        """Intenta usar textract como método universal"""
        if not TEXTRACT_AVAILABLE:
            raise ImportError("textract no disponible")
        
        text = textract.process(filepath).decode('utf-8')
        return text
    
    def _try_pypandoc(self, filepath: str) -> str:
        """Intenta usar pypandoc para conversión"""
        if not PYPANDOC_AVAILABLE:
            raise ImportError("pypandoc no disponible")
        
        # Crear archivo temporal de salida
        with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            pypandoc.convert_file(filepath, 'plain', outputfile=tmp_path)
            with open(tmp_path, 'r', encoding='utf-8') as f:
                return f.read()
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def _try_raw_text(self, filepath: str) -> str:
        """Último recurso: leer como texto crudo"""
        with open(filepath, 'rb') as f:
            raw_bytes = f.read()
        
        # Intentar decodificar con diferentes encodings
        encodings = ['utf-8', 'latin-1', 'cp1252', 'ascii']
        
        for encoding in encodings:
            try:
                text = raw_bytes.decode(encoding, errors='ignore')
                # Limpiar caracteres no imprimibles
                clean_text = ''.join(char for char in text if char.isprintable() or char.isspace())
                if len(clean_text) > 50:  # Solo si tiene contenido mínimo
                    return clean_text
            except:
                continue
        
        return ""

def extract_sections_mejorado(text: str) -> Dict[str, str]:
    """Extrae secciones legales con patrones mejorados."""
    sections = {"hechos": "", "derecho": "", "petitorio": "", "prueba": ""}
    
    # Patrones más robustos para detectar secciones
    patterns = {
        "hechos": [
            r"(?i)(hechos?|i\s*[-.)]\s*hechos?|primero|1\s*[-.)]\s*hechos?)",
            r"(?i)(antecedentes|relación de hechos|exposición de hechos)"
        ],
        "derecho": [
            r"(?i)(derecho|ii\s*[-.)]\s*derecho|segundo|2\s*[-.)]\s*derecho)",
            r"(?i)(fundamentos? de derecho|marco jurídico|base legal)"
        ],
        "petitorio": [
            r"(?i)(petitorio|solicito|pido|iii\s*[-.)]\s*petitorio|tercero|3\s*[-.)]\s*petitorio)",
            r"(?i)(se pida|ruego|solicita)"
        ],
        "prueba": [
            r"(?i)(prueba|ofrezco|iv\s*[-.)]\s*prueba|cuarto|4\s*[-.)]\s*prueba)",
            r"(?i)(ofrecimiento de prueba|medios probatorios)"
        ]
    }
    
    lines = text.split('\n')
    current_section = None
    
    for line in lines:
        line_clean = line.strip()
        if not line_clean:
            continue
        
        # Detectar cambio de sección
        section_found = None
        for section, pattern_list in patterns.items():
            for pattern in pattern_list:
                if re.search(pattern, line_clean):
                    section_found = section
                    break
            if section_found:
                break
        
        if section_found:
            current_section = section_found
        elif current_section:
            sections[current_section] += line_clean + " "
    
    # Si no se encontraron secciones, poner todo en hechos SIN LIMITACIONES
    if not any(sections.values()):
        sections["hechos"] = text  # Preservar contenido completo
    
    return sections

def procesar_archivo_universal(filepath: str, output_folder: str = "data/casos_anteriores") -> bool:
    """Procesa un archivo de cualquier tipo soportado."""
    processor = ProcessorUniversal()
    
    try:
        # Extraer texto
        text, method = processor.extract_text_from_file(filepath)
        
        if not text.strip():
            print(f"  ❌ No se pudo extraer texto válido")
            return False
        
        # Extraer secciones legales
        sections = extract_sections_mejorado(text)
        
        # Inferir tipo por carpeta
        carpeta = os.path.basename(os.path.dirname(filepath)).lower()
        tipos_mapeados = {
            "despido": "Despido Injustificado",
            "discriminatorio": "Despido Discriminatorio",
            "accidente": "Accidente Laboral",
            "licencia": "Licencia Médica",
            "solidaridad": "Solidaridad Laboral",
            "blanco": "Empleados en Blanco",
            "negro": "Empleados en Negro"
        }
        
        tipo_inferido = "Caso Legal"
        for key, tipo in tipos_mapeados.items():
            if key in carpeta:
                tipo_inferido = tipo
                break
        
        # Crear JSON
        filename = Path(filepath).stem
        json_data = {
            "id": filename,
            "tipo": tipo_inferido,
            "jurisdiccion": "Argentina",
            "empresa": "",
            "fecha": "",
            **sections,
            "archivo_origen": filepath,
            "metodo_extraccion": method,
            "extension_original": Path(filepath).suffix,
            "procesado_exitosamente": True
        }
        
        # Guardar JSON
        os.makedirs(output_folder, exist_ok=True)
        output_path = os.path.join(output_folder, f"{filename}.json")
        
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        
        print(f"  ✅ Guardado → {tipo_inferido}")
        return True
        
    except Exception as e:
        print(f"  ❌ Error: {str(e)[:100]}")
        return False

def procesar_todos_universal():
    """Procesa todos los archivos soportados en ejemplos_demandas/"""
    print("🚀 PROCESADOR UNIVERSAL DE DOCUMENTOS LEGALES")
    print("=" * 60)
    
    processor = ProcessorUniversal()
    
    # Buscar todos los archivos soportados
    archivos_encontrados = []
    for root, dirs, files in os.walk("ejemplos_demandas"):
        for file in files:
            file_path = os.path.join(root, file)
            file_ext = Path(file).suffix.lower()
            
            if file_ext in processor.supported_formats:
                archivos_encontrados.append(file_path)
    
    if not archivos_encontrados:
        print("⚠️ No se encontraron archivos soportados")
        return
    
    # Estadísticas por tipo
    stats_por_tipo = {}
    for archivo in archivos_encontrados:
        ext = Path(archivo).suffix.lower()
        stats_por_tipo[ext] = stats_por_tipo.get(ext, 0) + 1
    
    print(f"📂 Archivos encontrados: {len(archivos_encontrados)}")
    for ext, count in stats_por_tipo.items():
        print(f"   • {ext}: {count} archivos")
    print()
    
    # Procesar archivos
    exitosos = 0
    errores = []
    
    for archivo in archivos_encontrados:
        if procesar_archivo_universal(archivo):
            exitosos += 1
        else:
            errores.append(os.path.basename(archivo))
    
    # Reporte final
    print("\n" + "=" * 60)
    print("📊 REPORTE FINAL")
    print(f"✅ Exitosos: {exitosos}/{len(archivos_encontrados)}")
    print(f"❌ Errores: {len(errores)}")
    
    if errores:
        print("\n❌ Archivos problemáticos:")
        for error in errores[:10]:
            print(f"   • {error}")
        if len(errores) > 10:
            print(f"   ... y {len(errores) - 10} más")
    
    print(f"\n📁 Archivos JSON guardados en: data/casos_anteriores/")
    return exitosos, len(errores)

if __name__ == "__main__":
    procesar_todos_universal() 